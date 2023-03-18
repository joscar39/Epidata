import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service


# class TestLambda(unittest.TestCase):

    # def setUp(self):
    # chrome_options = Options()
    # chrome_options.add_argument("--headless")
    # self.driver = webdriver.Chrome(
    #     chrome_options=chrome_options, executable_path=r"C:\Users\user\Documents"
    #                                                    r"\Habits.ai\Automatizacion\q"
    #                                                    r"a_automation\recursos\chromedriver.exe")
    # def setUp(self):
path = Service(r"C:\Users\user\Documents\Epidata\src\Resources\chromedriver.exe")
driver = webdriver.Chrome(service=path)

    # Eliminar todos los retos para la empresa
    # def test_case1(self):
web = "https://lambdatest.github.io/sample-todo-app/"
# driver = self.driver
driver.get(web)
driver.maximize_window()
filesheet = r"C:\Users\user\Documents\Epidata\src\Resources\checks.xlsx"
wb = openpyxl.load_workbook(filesheet)
datos = wb["Hoja1"]
name = datos['A1']
first_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//body/div[1]/div[1]/div[1]/ul[1]/li[1]/input[1]")))
first_check.click()
driver.find_element(By.XPATH, "//body/div[1]/div[1]/div[1]/ul[1]/li[2]/input[1]").click()
time.sleep(1)
driver.find_element(By.XPATH, "//input[@id='sampletodotext']").send_keys(name.value)
time.sleep(1)
driver.find_element(By.XPATH, "//input[@id='addbutton']").click()
time.sleep(1)
result_test = False
search_new_check = driver.find_elements(By.CLASS_NAME, "done-false")
for search in search_new_check:
    if search.text in name.value:
        result_test = True
    else:
        result_test = False
assert result_test is True, "Error no se encontro el elemento"
driver.get_screenshot_as_file(
                f"C:\\Users\\user\\Documents\\Epidata\\src\\Resources\\screenshot\\se muestra el check de nombre {name.value}.png")
print(f"Se completo la prueba exitosamente, para el registro {name.value}")
wb.close()
# Load the Excel file
wb = openpyxl.load_workbook(filesheet)
# Select the sheet to work with
ws = wb['Hoja1']
# Get the data for the first user
primera_fila = ws[1]
# Delete the first row
ws.delete_rows(1)
# Add the first user data to the last row
ws.append(primera_fila)
# Save the updated Excel file
wb.save(filesheet)

# def tearDown(self):
driver.close()
#
#
# if __name__ == '__main__':
#     suite = unittest.TestLoader().loadTestsFromTestCase(TestLambda)
#     unittest.TextTestRunner(verbosity=2).run(suite)
